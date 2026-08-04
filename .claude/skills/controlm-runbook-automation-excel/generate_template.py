"""Render template-spec.yaml -> controlm-runbook-template.xlsx (the 2-tab
minimum-viable Control-M application runbook).

Run from the repo root:

    poetry run python .claude/skills/controlm-runbook-automation-excel/generate_template.py

The spec is the source of truth — edit it, then regenerate. The workbook is a
committed ASSET, not a drift-guarded render (xlsx zips embed timestamps, so
byte-identity across runs is not promised; content identity is).

Color convention (carried over from the source workbook): rows/columns whose
`source:` is `manual` or `graph-partial` are tinted — yellow means "a human
must capture or confirm this"; untinted means the graph fills it.
"""

from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
SPEC = HERE / "template-spec.yaml"
OUT = HERE / "controlm-runbook-template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="FFFF00")  # source workbook: yellow headers
MANUAL_FILL = PatternFill("solid", fgColor="FFF2CC")  # needs human capture
PARTIAL_FILL = PatternFill("solid", fgColor="FCE4B0")  # seam exists, needs enrichment
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _fill_for(source: str) -> PatternFill | None:
    if source == "manual":
        return MANUAL_FILL
    if source == "graph-partial":
        return PARTIAL_FILL
    return None


def _technical_details(ws, tab: dict) -> None:
    ws.append(tab["columns"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
    ws.append(["", "", tab["legend"], ""])
    for row in tab["rows"]:
        ws.append([row["info"], row.get("example", ""), row.get("comment", ""), ""])
        fill = _fill_for(row["source"])
        for cell in ws[ws.max_row]:
            cell.border = BORDER
            cell.alignment = WRAP
            if fill:
                cell.fill = fill
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 10
    ws.freeze_panes = "A2"


def _job_details(ws, tab: dict) -> None:
    cols = tab["columns"]
    ws.append([c["name"] for c in cols])
    for i, (cell, col) in enumerate(zip(ws[1], cols, strict=True), start=1):
        cell.font = Font(bold=True)
        cell.fill = _fill_for(col["source"]) or HEADER_FILL
        cell.border = BORDER
        cell.alignment = WRAP
        ws.column_dimensions[get_column_letter(i)].width = max(18, min(46, len(col["name"]) + 4))
    for values in tab.get("example_rows", []):
        assert len(values) == len(
            cols
        ), f"example row has {len(values)} cells, spec has {len(cols)} columns"
        ws.append(values)
        for cell in ws[ws.max_row]:
            cell.border = BORDER
            cell.alignment = WRAP
    ws.freeze_panes = "C2"


def main() -> None:
    spec = yaml.safe_load(SPEC.read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)
    for tab in spec["tabs"]:
        ws = wb.create_sheet(tab["name"])
        if tab["layout"] == "rows":
            _technical_details(ws, tab)
        else:
            _job_details(ws, tab)
    wb.save(OUT)
    print(f"wrote {OUT} ({len(spec['tabs'])} tabs)")


if __name__ == "__main__":
    main()
